from __future__ import annotations
import re
from io import BytesIO
from decimal import Decimal
from datetime import date, datetime
from typing import Any
from sqlalchemy import select
from datetime import datetime
from fastapi import UploadFile, File
from fastapi.responses import JSONResponse
from app.excel_repair import repair_xlsx_bytes

import openpyxl

from datetime import date, timedelta
from typing import Any

from fastapi import APIRouter, Request
from fastapi.responses import JSONResponse

from app.db import SessionLocal
from app.ozon_performance_api import get_perf_client, OzonPerformanceApiError
from decimal import Decimal, InvalidOperation
from datetime import datetime
from sqlalchemy import select
from app.models import Product, TrackedCampaign, AdsTopDaily, AdsStencilDaily, AdsTotalDaily, ConversionsDaily

router = APIRouter()

PLACEMENT_TOP = "PLACEMENT_TOP_PROMOTION"
PLACEMENT_TRAF_SET = {
    # у тебя в примере "Трафареты" имеют именно это:
    "PLACEMENT_SEARCH_AND_CATEGORY",
    # на всякий случай (если Ozоn отдаст вариации):
    "PLACEMENT_SEARCH_PROMOTION",
    "PLACEMENT_STENCIL",
    "PLACEMENT_TRAFFIC",
}

BID_SCALE = Decimal("1000000")  # Ozon bid приходит в micro-рублях: 47_000_000 => 47.00 ₽


def _bid_to_rub(v: Any) -> Decimal | None:
    """
    Переводит bid из micro-рублей в рубли (Decimal с 2 знаками).
    None -> None
    """
    if v is None:
        return None
    d = _parse_ru_decimal(v)          # превращаем строку/число в Decimal
    return (d / BID_SCALE).quantize(Decimal("0.01"))


def _has_any_placement(c: dict[str, Any], need: str | set[str]) -> bool:
    placements = c.get("placement") or []
    if isinstance(need, set):
        return any(p in need for p in placements)
    return need in placements


async def _get_active_product_and_sku(request: Request) -> tuple[int, int, str]:
    """
    Возвращает:
      (ozon_account_id, product_id, sku_str)
    Берём из session:
      active_ozon_account_id (у тебя уже используется)
      active_product_id      (выбранный товар на РНП)
    Проверяем, что товар принадлежит выбранному кабинету.
    """
    ozon_account_id = request.session.get("active_ozon_account_id")
    if not ozon_account_id:
        raise OzonPerformanceApiError("Не выбран Ozon-кабинет (active_ozon_account_id отсутствует)")

    product_id = request.session.get("active_product_id")
    if not product_id:
        raise OzonPerformanceApiError("Не выбран товар на РНП (active_product_id отсутствует)")

    async with SessionLocal() as session:
        product = await session.get(Product, int(product_id))

    if not product:
        raise OzonPerformanceApiError(f"Товар product_id={product_id} не найден в БД")

    if int(product.ozon_account_id) != int(ozon_account_id):
        raise OzonPerformanceApiError(
            f"Товар product_id={product_id} не принадлежит выбранному кабинету ozon_account_id={ozon_account_id}"
        )

    if product.sku is None:
        raise OzonPerformanceApiError(
            f"У товара product_id={product_id} не заполнено поле sku — не с чем матчить кампании"
        )

    return int(ozon_account_id), int(product.product_id), str(int(product.sku))


async def _get_tracked_campaign_ids(ozon_account_id: int) -> set[str]:
    async with SessionLocal() as session:
        rows = await session.execute(
            select(TrackedCampaign.campaign_id).where(
                TrackedCampaign.ozon_account_id == int(ozon_account_id)
            )
        )
        # rows.scalars().all() -> list[str]
        return {str(x) for x in rows.scalars().all() if x}


async def _collect_campaign_match(api, campaign: dict[str, Any], target_sku: str) -> dict[str, Any]:
    """
    Находит target_sku среди товаров кампании.
    Если найдено:
      - ours_bid: bid из campaign_products_v2_all
      - competitor_bid: bid из competitive endpoint
    """
    campaign_id = str(campaign.get("id"))
    title = campaign.get("title")
    placements = campaign.get("placement") or []

    products = await api.campaign_products_v2_all(campaign_id=campaign_id, page_size=200)

    matched: dict[str, Any] | None = None
    for p in products:
        sku = p.get("sku")
        if sku is None:
            continue
        if str(sku) == str(target_sku):
            matched = p
            break

    if not matched:
        return {
            "campaignId": campaign_id,
            "title": title,
            "placement": placements,
            "matched": False,
            "target_sku": target_sku,
        }

    ours_bid = matched.get("bid")  # как у тебя: "22000000"

    competitive = await api.campaign_products_bids_competitive_all(
        campaign_id=campaign_id,
        skus=[target_sku],
        chunk_size=200,
    )
    competitor_bid = None
    bids = competitive.get("bids") or []
    if bids:
        competitor_bid = (bids[0] or {}).get("bid")  # как у тебя: "46000000"

    return {
        "campaignId": campaign_id,
        "title": title,
        "placement": placements,
        "matched": True,
        "target_sku": target_sku,
        "ours_bid": ours_bid,
        "competitor_bid": competitor_bid,
        # пока для проверки оставим “сырые” куски:
        "raw_product": matched,
        "raw_competitive": competitive,
    }


def _parse_ru_decimal(v: Any) -> Decimal:
    """
    Преобразует строки вида '4 500,00' / '0,00' / '9000,00' / '0.00' в Decimal.
    Пустое/None -> Decimal(0)
    """
    if v is None:
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))

    s = str(v).strip()
    if not s:
        return Decimal("0")

    # убрать валюты/пробелы
    s = s.replace("₽", "").replace("\u00a0", " ").replace(" ", "")
    # русская запятая -> точка
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _parse_int(v: Any) -> int:
    if v is None:
        return 0
    try:
        s = str(v).strip().replace(" ", "")
        if not s:
            return 0
        # иногда может быть '0,00' — тогда int(Decimal)
        if "," in s or "." in s:
            return int(_parse_ru_decimal(s))
        return int(s)
    except Exception:
        return 0


def _parse_ctr_percent_to_ratio(v: Any) -> Decimal | None:
    """
    В stats может прийти ctr '0,00' — это проценты.
    В БД ctr у тебя Numeric(10,4) без строгого определения.
    Я предлагаю хранить как проценты (0..100), чтобы совпадало с API.
    Если хочешь хранить как долю (0..1) — скажи, поменяю.
    """
    if v is None:
        return None
    d = _parse_ru_decimal(v)
    return d


def _safe_div(a: Decimal, b: Decimal) -> Decimal | None:
    if b is None or b == 0:
        return None
    return a / b


def _max_decimal(values: list[Any]) -> Decimal | None:
    ds: list[Decimal] = []
    for x in values:
        if x is None:
            continue
        d = _parse_ru_decimal(x)
        ds.append(d)
    return max(ds) if ds else None


def _aggregate_stats_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """
    Агрегация по placement (если несколько кампаний одного типа).
    Суммируем spend/orders/clicks/views/toCart/ordersMoney.
    CTR/DRR/clickPrice — берём средневзвешенно по кликам/спенду (упрощённо) или max?:
      - ctr: (sum clicks / sum views) * 100 (если views>0)
      - drr: (sum spend / sum ordersMoney) * 100 (если ordersMoney>0)
      - clickPrice: sum spend / sum clicks (если clicks>0)
    """
    spend = sum((_parse_ru_decimal(r.get("moneySpent")) for r in rows), Decimal("0"))
    orders_money = sum((_parse_ru_decimal(r.get("ordersMoney")) for r in rows), Decimal("0"))
    orders_cnt = sum((_parse_int(r.get("orders")) for r in rows))
    views = sum((_parse_int(r.get("views")) for r in rows))
    clicks = sum((_parse_int(r.get("clicks")) for r in rows))
    carts = sum((_parse_int(r.get("toCart")) for r in rows))

    ctr = None
    if views > 0:
        ctr = (Decimal(clicks) / Decimal(views)) * Decimal("100")

    drr = None
    if orders_money > 0:
        drr = (spend / orders_money) * Decimal("100")

    click_price = None
    if clicks > 0:
        click_price = spend / Decimal(clicks)

    return {
        "spend_rub": spend,
        "orders_amount_rub": orders_money,
        "orders_count": orders_cnt,
        "impressions": views,
        "clicks": clicks,
        "carts": carts,
        "ctr": ctr,
        "drr": drr,
        "click_price": click_price,
    }


async def _upsert_ads_top_daily(
    *,
    session,
    ozon_account_id: int,
    product_id: int,
    day_obj,
    agg: dict[str, Any],
    bid_ours: Decimal | None,
    bid_comp: Decimal | None,
) -> None:
    q = await session.execute(
        select(AdsTopDaily).where(
            AdsTopDaily.ozon_account_id == ozon_account_id,
            AdsTopDaily.product_id == product_id,
            AdsTopDaily.day == day_obj,
        )
    )
    obj = q.scalar_one_or_none()
    if obj is None:
        obj = AdsTopDaily(
            ozon_account_id=ozon_account_id,
            product_id=product_id,
            day=day_obj,
        )
        session.add(obj)

    obj.spend_rub = agg["spend_rub"]
    obj.orders_amount_rub = agg["orders_amount_rub"]
    obj.orders_count = agg["orders_count"]
    obj.impressions = agg["impressions"]
    obj.clicks = agg["clicks"]
    obj.carts = agg["carts"]
    obj.ctr = agg["ctr"]
    obj.drr = agg["drr"]

    if bid_ours is not None:
        obj.bid_ours = bid_ours
    if bid_comp is not None:
        obj.bid_competitor = bid_comp

    # производные
    if agg["orders_count"] > 0:
        obj.spend_per_order = _safe_div(agg["spend_rub"], Decimal(agg["orders_count"]))
    else:
        obj.spend_per_order = None

    if agg["clicks"] > 0:
        obj.ad_conversion = _safe_div(Decimal(agg["carts"]) * Decimal("100"), Decimal(agg["clicks"]))
    else:
        obj.ad_conversion = Decimal("0")

    obj.updated_at = datetime.utcnow()



async def _upsert_ads_stencil_daily(
    *,
    session,
    ozon_account_id: int,
    product_id: int,
    day_obj,
    agg: dict[str, Any],
    bid_ours: Decimal | None,
    bid_comp: Decimal | None,
) -> None:
    q = await session.execute(
        select(AdsStencilDaily).where(
            AdsStencilDaily.ozon_account_id == ozon_account_id,
            AdsStencilDaily.product_id == product_id,
            AdsStencilDaily.day == day_obj,
        )
    )
    obj = q.scalar_one_or_none()
    if obj is None:
        obj = AdsStencilDaily(
            ozon_account_id=ozon_account_id,
            product_id=product_id,
            day=day_obj,
        )
        session.add(obj)

    obj.spend_rub = agg["spend_rub"]
    obj.orders_amount_rub = agg["orders_amount_rub"]
    obj.orders_count = agg["orders_count"]
    obj.impressions = agg["impressions"]
    obj.clicks = agg["clicks"]
    obj.carts = agg["carts"]
    obj.ctr = agg["ctr"]
    obj.drr = agg["drr"]

    if bid_ours is not None:
        obj.bid_ours = bid_ours
    if bid_comp is not None:
        obj.bid_competitor = bid_comp

    if agg["orders_count"] > 0:
        obj.spend_per_order = _safe_div(agg["spend_rub"], Decimal(agg["orders_count"]))
    else:
        obj.spend_per_order = None

    if agg["clicks"] > 0:
        obj.ad_conversion = _safe_div(Decimal(agg["carts"]) * Decimal("100"), Decimal(agg["clicks"]))
    else:
        obj.ad_conversion = Decimal("0")

    obj.updated_at = datetime.utcnow()


# ============================================================
# ✅ UPSERT ТОЛЬКО СТАВОК (для СЕГОДНЯ)
# ============================================================

async def _upsert_ads_top_bids_only(
    *,
    session,
    ozon_account_id: int,
    product_id: int,
    day_obj,
    bid_ours: Decimal | None,
    bid_comp: Decimal | None,
) -> None:
    q = await session.execute(
        select(AdsTopDaily).where(
            AdsTopDaily.ozon_account_id == ozon_account_id,
            AdsTopDaily.product_id == product_id,
            AdsTopDaily.day == day_obj,
        )
    )
    obj = q.scalar_one_or_none()

    if obj is None:
        # ⚠️ создаём строку на today, если её ещё нет
        obj = AdsTopDaily(
            ozon_account_id=ozon_account_id,
            product_id=product_id,
            day=day_obj,
        )
        session.add(obj)

    if bid_ours is not None:
        obj.bid_ours = bid_ours
    if bid_comp is not None:
        obj.bid_competitor = bid_comp

    obj.updated_at = datetime.utcnow()


async def _upsert_ads_stencil_bids_only(
    *,
    session,
    ozon_account_id: int,
    product_id: int,
    day_obj,
    bid_ours: Decimal | None,
    bid_comp: Decimal | None,
) -> None:
    q = await session.execute(
        select(AdsStencilDaily).where(
            AdsStencilDaily.ozon_account_id == ozon_account_id,
            AdsStencilDaily.product_id == product_id,
            AdsStencilDaily.day == day_obj,
        )
    )
    obj = q.scalar_one_or_none()

    if obj is None:
        obj = AdsStencilDaily(
            ozon_account_id=ozon_account_id,
            product_id=product_id,
            day=day_obj,
        )
        session.add(obj)

    if bid_ours is not None:
        obj.bid_ours = bid_ours
    if bid_comp is not None:
        obj.bid_competitor = bid_comp

    obj.updated_at = datetime.utcnow()


@router.post("/api/rnp/import-yesterday")
async def rnp_import_yesterday(
    request: Request,
    promo_report: UploadFile = File(...),   # ✅ отчёт "Аналитика продаж/продвижения" (то, что уже работало)
    sales_report: UploadFile = File(...),   # ✅ отчёт "Продажи товара" (новый, битый)
):
    try:
        api = await get_perf_client(request)
        ozon_account_id, product_id, target_sku = await _get_active_product_and_sku(request)

        today_obj = date.today()
        yesterday_obj = today_obj - timedelta(days=1)
        day = yesterday_obj.strftime("%Y-%m-%d")
        today_iso = today_obj.strftime("%Y-%m-%d")

        # =========================
        # 1) ТОП/ТРАФ как раньше (твой код без изменений по сути)
        # =========================
        campaigns_data = await api.campaign_list()
        campaigns_all = campaigns_data.get("list", []) or []

        tracked_ids = await _get_tracked_campaign_ids(ozon_account_id)
        campaigns = [c for c in campaigns_all if str(c.get("id")) in tracked_ids]

        top_campaigns = [c for c in campaigns if _has_any_placement(c, PLACEMENT_TOP)]
        traf_campaigns = [c for c in campaigns if _has_any_placement(c, PLACEMENT_TRAF_SET)]

        top_results = [await _collect_campaign_match(api, c, target_sku) for c in top_campaigns]
        traf_results = [await _collect_campaign_match(api, c, target_sku) for c in traf_campaigns]

        campaign_ids = [str(c["id"]) for c in campaigns if c.get("id")]
        stats = await api.statistics_campaign_product(
            campaign_ids=campaign_ids,
            date_from=day,
            date_to=day,
        )

        stats_rows = (stats.get("rows") or []) if isinstance(stats, dict) else []
        used_rows = [r for r in stats_rows if str(r.get("id")) in tracked_ids]

        top_rows = [r for r in used_rows if (r.get("placement") == "top-promotion")]
        traf_rows = [r for r in used_rows if (r.get("placement") == "search-and-category")]

        top_agg = _aggregate_stats_rows(top_rows) if top_rows else None
        traf_agg = _aggregate_stats_rows(traf_rows) if traf_rows else None

        top_bid_ours = _bid_to_rub(_max_decimal([x.get("ours_bid") for x in top_results if x.get("matched")]))
        top_bid_comp = _bid_to_rub(_max_decimal([x.get("competitor_bid") for x in top_results if x.get("matched")]))
        traf_bid_ours = _bid_to_rub(_max_decimal([x.get("ours_bid") for x in traf_results if x.get("matched")]))
        traf_bid_comp = _bid_to_rub(_max_decimal([x.get("competitor_bid") for x in traf_results if x.get("matched")]))

        # =========================
        # 2) ИТОГО из promo_report (как раньше)
        # =========================
        promo_bytes = await promo_report.read()
        total_parsed = parse_total_from_promo_analytics_xlsx(promo_bytes, target_sku)
        total_day_obj = total_parsed["day_obj"] or yesterday_obj

        # =========================
        # 2.1) Новый файл sales_report: сначала чиним, дальше ПОКА НЕ ТРОГАЕМ логику
        # (сейчас только проверяем что после починки он открывается)
        # =========================
        sales_bytes = await sales_report.read()
        sales_fixed_bytes = repair_xlsx_bytes(sales_bytes)

        # ✅ читаем все заказы + конверсии из нового отчёта
        sales_parsed = parse_orders_and_conversions_from_sales_report_xlsx(
            sales_fixed_bytes,
            target_sku=target_sku,  # ✅ матч по SKU
        )

        sales_day_obj = sales_parsed["day_obj"]
        if sales_day_obj and (total_day_obj != sales_day_obj):
            # на всякий случай — но чаще всего совпадает
            total_day_obj = sales_day_obj

        all_orders_amount_rub = sales_parsed["all_orders_amount_rub"].quantize(Decimal("0.01"))
        all_orders_count = int(sales_parsed["all_orders_count"])

        drr_total = None
        if all_orders_amount_rub and all_orders_amount_rub > 0 and total_parsed["spend_rub"] is not None:
            drr_total = (total_parsed["spend_rub"] / all_orders_amount_rub) * Decimal("100")

        total_agg = {
            "spend_rub": total_parsed["spend_rub"].quantize(Decimal("0.01")),
            "ad_orders_amount_rub": total_parsed["ad_orders_amount_rub"].quantize(Decimal("0.01")),
            "ad_orders_count": int(total_parsed["ad_orders_count"]),
            "impressions": int(total_parsed["impressions"]),
            "clicks": int(total_parsed["clicks"]),
            "carts": int(total_parsed["carts"]),
            "drr": (total_parsed["drr_weighted"].quantize(Decimal("0.0001")) if total_parsed["drr_weighted"] is not None else None),
            "ctr": (total_parsed["ctr_weighted"].quantize(Decimal("0.0001")) if total_parsed["ctr_weighted"] is not None else None),
            "ad_conversion": (total_parsed["ad_conversion_weighted"].quantize(Decimal("0.0001")) if total_parsed["ad_conversion_weighted"] is not None else None),
            "all_orders_amount_rub": all_orders_amount_rub,
            "all_orders_count": all_orders_count,
            "drr_total": (drr_total.quantize(Decimal("0.0001")) if drr_total is not None else None),
        }

        # =========================
        # 3) Сохранение в БД
        # =========================
        day_obj = yesterday_obj
        bid_day_obj = today_obj

        async with SessionLocal() as session:
            # ТОП/ТРАФ метрики (вчера)
            if top_agg is not None:
                await _upsert_ads_top_daily(
                    session=session,
                    ozon_account_id=ozon_account_id,
                    product_id=product_id,
                    day_obj=day_obj,
                    agg=top_agg,
                    bid_ours=None,
                    bid_comp=None,
                )
            if traf_agg is not None:
                await _upsert_ads_stencil_daily(
                    session=session,
                    ozon_account_id=ozon_account_id,
                    product_id=product_id,
                    day_obj=day_obj,
                    agg=traf_agg,
                    bid_ours=None,
                    bid_comp=None,
                )

            # ставки (сегодня)
            if top_bid_ours is not None or top_bid_comp is not None:
                await _upsert_ads_top_bids_only(
                    session=session,
                    ozon_account_id=ozon_account_id,
                    product_id=product_id,
                    day_obj=bid_day_obj,
                    bid_ours=top_bid_ours,
                    bid_comp=top_bid_comp,
                )
            if traf_bid_ours is not None or traf_bid_comp is not None:
                await _upsert_ads_stencil_bids_only(
                    session=session,
                    ozon_account_id=ozon_account_id,
                    product_id=product_id,
                    day_obj=bid_day_obj,
                    bid_ours=traf_bid_ours,
                    bid_comp=traf_bid_comp,
                )

            # ✅ ИТОГО (из файла)
            await _upsert_ads_total_daily_from_report(
                session=session,
                ozon_account_id=ozon_account_id,
                product_id=product_id,
                day_obj=total_day_obj,
                agg=total_agg,
            )

            # ✅ КОНВЕРСИИ (из sales-отчёта)
            await _upsert_conversions_daily_from_report(
                session=session,
                ozon_account_id=ozon_account_id,
                product_id=product_id,
                day_obj=total_day_obj,
                conv=sales_parsed["conversions"],
            )

            await session.commit()

        # ✅ Возвращаем короткий ответ (без простыни JSON)
        return JSONResponse(content={
            "ok": True,
            "message": "Импорт выполнен: ТОП/ТРАФ из API, ИТОГО из отчёта.",
            "day_api": day,
            "day_total": total_day_obj.isoformat(),
            "total_campaigns_in_file": total_parsed["campaigns_in_file_for_sku"],
            "saved": {
                "top_yesterday": bool(top_rows),
                "traf_yesterday": bool(traf_rows),
                "total_from_file": True,
                "bids_today_top": bool(top_bid_ours is not None or top_bid_comp is not None),
                "bids_today_traf": bool(traf_bid_ours is not None or traf_bid_comp is not None),
            }
        })

    except OzonPerformanceApiError as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": f"Unexpected error: {e}"})


def _parse_period_from_sheet(ws) -> date | None:
    """
    В A1 обычно: 'Период: 07.02.2026 - 07.02.2026'
    Берём первую дату как day.
    """
    v = ws.cell(1, 1).value
    if not v:
        return None
    s = str(v)
    m = re.search(r"(\d{2}\.\d{2}\.\d{4})", s)
    if not m:
        return None
    return datetime.strptime(m.group(1), "%d.%m.%Y").date()


def _find_header_row(ws, needle: str = "SKU", max_scan_rows: int = 50) -> int | None:
    for r in range(1, max_scan_rows + 1):
        v = ws.cell(r, 1).value
        if v and str(v).strip() == needle:
            return r
    # иногда SKU может быть не в 1 колонке — на всякий случай по всей строке
    for r in range(1, max_scan_rows + 1):
        for c in range(1, 40):
            v = ws.cell(r, c).value
            if v and str(v).strip() == needle:
                return r
    return None


def _row_as_dict(ws, row_idx: int, col_map: dict[str, int]) -> dict[str, Any]:
    d: dict[str, Any] = {}
    for k, c in col_map.items():
        d[k] = ws.cell(row_idx, c).value
    return d


def _to_decimal(v: Any) -> Decimal:
    # используй твою _parse_ru_decimal — просто переиспользуем
    return _parse_ru_decimal(v)


def _to_int(v: Any) -> int:
    return _parse_int(v)


def _avg_decimal(values: list[Decimal | None]) -> Decimal | None:
    xs = [x for x in values if x is not None]
    if not xs:
        return None
    return sum(xs, Decimal("0")) / Decimal(len(xs))


def _is_empty_cell(v: Any) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s == "-"


def parse_total_from_promo_analytics_xlsx(file_bytes: bytes, target_sku: str) -> dict[str, Any]:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    if "Statistics" not in wb.sheetnames:
        raise ValueError("В отчёте нет листа 'Statistics'")

    ws = wb["Statistics"]

    day_obj = _parse_period_from_sheet(ws)

    header_row = _find_header_row(ws, "SKU")
    if not header_row:
        raise ValueError("Не нашёл строку заголовков (SKU) на листе Statistics")

    # читаем заголовки -> индекс колонки
    header_to_col: dict[str, int] = {}
    for c in range(1, 60):
        name = ws.cell(header_row, c).value
        if not name:
            continue
        header_to_col[str(name).strip()] = c

    # обязательные колонки
    need = [
        "SKU",
        "ID кампании",
        "Расход, ₽",
        "ДРР, %",
        "Продажи, ₽",
        "Заказы, шт",
        "CTR, %",
        "Показы",
        "Клики",
        "В корзину",
    ]
    for n in need:
        if n not in header_to_col:
            raise ValueError(f"В отчёте нет колонки: {n}")

    col_map = {n: header_to_col[n] for n in need}

    # собираем строки по SKU
    rows: list[dict[str, Any]] = []
    r = header_row + 1
    while True:
        sku_val = ws.cell(r, col_map["SKU"]).value
        if sku_val is None or str(sku_val).strip() == "":
            break  # конец таблицы
        if str(int(sku_val)) == str(target_sku):
            rows.append(_row_as_dict(ws, r, col_map))
        r += 1

    if not rows:
        raise ValueError(f"В отчёте нет строк по SKU={target_sku}")

    # агрегаты суммой
    spend = sum((_to_decimal(x.get("Расход, ₽")) for x in rows), Decimal("0"))
    ad_orders_amount = sum((_to_decimal(x.get("Продажи, ₽")) for x in rows), Decimal("0"))
    ad_orders_count = sum((_to_int(x.get("Заказы, шт")) for x in rows))
    impressions = sum((_to_int(x.get("Показы")) for x in rows))
    clicks = sum((_to_int(x.get("Клики")) for x in rows))
    carts = sum((_to_int(x.get("В корзину")) for x in rows))

    # ------------------------------------------------------------
    # ✅ ИТОГО: ВЗВЕШЕННОЕ (как в интерфейсе)
    # Конверсия в корзину = carts/clicks * 100
    # CTR = clicks/impressions * 100
    # DRR = spend/sales * 100
    # ------------------------------------------------------------

    # Важный момент:
    # - если в ячейке "-" или пусто → строку НЕ учитываем для этой метрики
    # - если знаменатель 0 → метрика = None (неопределено)

    # для DRR
    spend_sum = Decimal("0")
    sales_sum = Decimal("0")

    # для CTR
    clicks_sum_for_ctr = Decimal("0")
    impressions_sum_for_ctr = Decimal("0")

    # для конверсии в корзину
    carts_sum_for_conv = Decimal("0")
    clicks_sum_for_conv = Decimal("0")

    for x in rows:
        # ----- DRR -----
        spend_raw = x.get("Расход, ₽")
        sales_raw = x.get("Продажи, ₽")
        if not _is_empty_cell(spend_raw):
            spend_sum += _to_decimal(spend_raw)
        if not _is_empty_cell(sales_raw):
            sales_sum += _to_decimal(sales_raw)

        # ----- CTR -----
        clk_raw = x.get("Клики")
        impr_raw = x.get("Показы")
        if not _is_empty_cell(clk_raw) and not _is_empty_cell(impr_raw):
            clk = Decimal(_to_int(clk_raw))
            impr = Decimal(_to_int(impr_raw))
            if impr > 0:
                clicks_sum_for_ctr += clk
                impressions_sum_for_ctr += impr

        # ----- CONV (клик -> корзина) -----
        cart_raw = x.get("В корзину")
        clk_raw2 = x.get("Клики")
        if not _is_empty_cell(cart_raw) and not _is_empty_cell(clk_raw2):
            cart = Decimal(_to_int(cart_raw))
            clk2 = Decimal(_to_int(clk_raw2))
            if clk2 > 0:
                carts_sum_for_conv += cart
                clicks_sum_for_conv += clk2

    drr_weighted = (spend_sum / sales_sum * Decimal("100")) if sales_sum > 0 else None
    ctr_weighted = (
                clicks_sum_for_ctr / impressions_sum_for_ctr * Decimal("100")) if impressions_sum_for_ctr > 0 else None
    conv_weighted = (carts_sum_for_conv / clicks_sum_for_conv * Decimal("100")) if clicks_sum_for_conv > 0 else None

    return {
        "day_obj": day_obj,  # может быть None — тогда возьмём yesterday
        "campaigns_in_file_for_sku": len(rows),
        "spend_rub": spend,
        "ad_orders_amount_rub": ad_orders_amount,
        "ad_orders_count": ad_orders_count,
        "impressions": impressions,
        "clicks": clicks,
        "carts": carts,
        "drr_weighted": drr_weighted,
        "ctr_weighted": ctr_weighted,
        "ad_conversion_weighted": conv_weighted,

    }


def _parse_period_from_sales_report(ws) -> date | None:
    """
    В sales-отчёте в A2: 'Период с 07.02.2026 по 07.02.2026'
    Берём первую дату.
    """
    v = ws.cell(2, 1).value
    if not v:
        return None
    m = re.search(r"(\d{2}\.\d{2}\.\d{4})", str(v))
    if not m:
        return None
    return datetime.strptime(m.group(1), "%d.%m.%Y").date()


def _parse_excel_percent_to_percent(v: Any) -> Decimal | None:
    """
    В этом отчёте проценты часто идут как доля (0.0575) при заголовке "%".
    Приводим к 0..100.
    """
    if v is None:
        return None

    # openpyxl часто отдаёт float
    d = _parse_ru_decimal(v)

    # если 0..1 — это доля, превращаем в %
    if Decimal("0") <= d <= Decimal("1"):
        d = d * Decimal("100")

    return d


def _norm_header(v: Any) -> str:
    if v is None:
        return ""
    s = str(v)
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def _find_header_row_by_title(ws, needle: str, max_scan_rows: int = 80, max_scan_cols: int = 80) -> int | None:
    n = _norm_header(needle)
    for r in range(1, max_scan_rows + 1):
        for c in range(1, max_scan_cols + 1):
            if _norm_header(ws.cell(r, c).value) == n:
                return r
    return None


def _build_sales_header_to_col(ws, header_row: int) -> dict[str, int]:
    """
    В этом отчёте:
      - на header_row (обычно 10) стоит SKU/Name/и т.п.
      - на header_row+1 (обычно 11) стоят метрики типа "Заказано на сумму", "Конверсия ..."

    Поэтому: если на header_row пусто, берём header_row+1.
    """
    header_to_col: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        top = ws.cell(header_row, c).value
        sub = ws.cell(header_row + 1, c).value
        name = sub if (sub is not None and str(sub).strip() != "") else top
        key = _norm_header(name)
        if key:
            header_to_col[key] = c
    return header_to_col


def _parse_excel_fraction_to_percent(v: Any) -> Decimal | None:
    """
    В sales-отчёте конверсии приходят как доли:
      0.0575 => 5.75%
    """
    if v is None:
        return None
    d = _parse_ru_decimal(v)
    # обычно 0..1, но если вдруг уже 0..100 — оставим как есть
    if Decimal("0") <= d <= Decimal("1"):
        d = d * Decimal("100")
    return d


def parse_orders_and_conversions_from_sales_report_xlsx(
    file_bytes: bytes,
    *,
    target_sku: str,
) -> dict[str, Any]:
    """
    Читает sales-отчёт (лист 'По товарам') и достаёт:
      - все заказы (руб/шт): "Заказано на сумму", "Заказано товаров"
      - конверсии (доли -> %)
    Матч по колонке 'SKU' == target_sku
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    sheet_name = "По товарам"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    day_obj = _parse_period_from_sales_report(ws)

    # ✅ заголовки ищем по SKU (а не Product ID)
    header_row = _find_header_row_by_title(ws, "SKU")
    if not header_row:
        raise ValueError("В sales-отчёте не нашёл строку заголовков (колонка 'SKU')")

    header_to_col = _build_sales_header_to_col(ws, header_row)

    # какие колонки нам нужны (в отчёте они БЕЗ символа % и часто с переносами строк)
    need_orders = {
        "sku": "sku",
        "заказано на сумму": "all_orders_amount_rub",
        "заказано товаров": "all_orders_count",
    }
    for k in need_orders.keys():
        if k not in header_to_col:
            raise ValueError(f"В sales-отчёте нет колонки: {k}")

    need_conv = {
        "конверсия из показа в заказ": "impression_to_order",
        "конверсия из поиска и каталога в корзину": "search_cat_to_cart",
        "конверсия из поиска и каталога в карточку": "search_cat_to_card",
        "конверсия из карточки в корзину": "card_to_cart",
        "конверсия в корзину общая": "cart_total",
        "конверсия из корзины в заказ": "cart_to_order",
        "конверсия из заказа в выкуп": "order_to_purchase",
    }

    for k in need_conv.keys():
        if k not in header_to_col:
            raise ValueError(f"В sales-отчёте нет колонки конверсии: {k}")

    col_sku = header_to_col["sku"]

    # ✅ данные начинаются ниже заголовков, в твоём файле это примерно с 14 строки,
    # но делаем универсально: идём вниз, пока SKU не закончится
    found_row = None
    r = header_row + 1
    empty_streak = 0
    while r <= ws.max_row:
        sku_val = ws.cell(r, col_sku).value
        if sku_val is None or str(sku_val).strip() == "":
            empty_streak += 1
            if empty_streak >= 30:
                break
            r += 1
            continue

        empty_streak = 0
        try:
            sku_int = str(int(str(sku_val).strip()))
        except Exception:
            r += 1
            continue

        if sku_int == str(int(target_sku)):
            found_row = r
            break

        r += 1

    if not found_row:
        raise ValueError(f"В sales-отчёте не нашёл строку по SKU={target_sku}")

    # --- ИТОГО: все заказы ---
    all_orders_amount = _parse_ru_decimal(ws.cell(found_row, header_to_col["заказано на сумму"]).value)
    all_orders_count  = _parse_int(ws.cell(found_row, header_to_col["заказано товаров"]).value)

    # --- КОНВЕРСИИ ---
    conv: dict[str, Decimal | None] = {}
    for title_norm, field in need_conv.items():
        v = ws.cell(found_row, header_to_col[title_norm]).value
        conv[field] = _parse_excel_fraction_to_percent(v)

    return {
        "day_obj": day_obj,
        "all_orders_amount_rub": all_orders_amount,
        "all_orders_count": all_orders_count,
        "conversions": conv,
    }


async def _upsert_ads_total_daily_from_report(
    *,
    session,
    ozon_account_id: int,
    product_id: int,
    day_obj: date,
    agg: dict[str, Any],
) -> None:
    q = await session.execute(
        select(AdsTotalDaily).where(
            AdsTotalDaily.ozon_account_id == ozon_account_id,
            AdsTotalDaily.product_id == product_id,
            AdsTotalDaily.day == day_obj,
        )
    )
    obj = q.scalar_one_or_none()
    if obj is None:
        obj = AdsTotalDaily(
            ozon_account_id=ozon_account_id,
            product_id=product_id,
            day=day_obj,
        )
        session.add(obj)

    # ✅ заполняем только то, что ты разрешил сейчас
    obj.spend_rub = agg["spend_rub"]
    obj.ad_orders_amount_rub = agg["ad_orders_amount_rub"]
    obj.ad_orders_count = agg["ad_orders_count"]
    obj.impressions = agg["impressions"]
    obj.clicks = agg["clicks"]
    obj.carts = agg["carts"]

    obj.drr = agg["drr"]
    obj.ctr = agg["ctr"]
    obj.ad_conversion = agg["ad_conversion"]

    # ✅ теперь заполняем ВСЕ заказы
    obj.all_orders_amount_rub = agg.get("all_orders_amount_rub")
    obj.all_orders_count = agg.get("all_orders_count")

    # ✅ общий DRR (от всех заказов)
    obj.drr_total = agg.get("drr_total")

    obj.updated_at = datetime.utcnow()


async def _upsert_conversions_daily_from_report(
    *,
    session,
    ozon_account_id: int,
    product_id: int,
    day_obj: date,
    conv: dict[str, Any],
) -> None:
    q = await session.execute(
        select(ConversionsDaily).where(
            ConversionsDaily.ozon_account_id == ozon_account_id,
            ConversionsDaily.product_id == product_id,
            ConversionsDaily.day == day_obj,
        )
    )
    obj = q.scalar_one_or_none()
    if obj is None:
        obj = ConversionsDaily(
            ozon_account_id=ozon_account_id,
            product_id=product_id,
            day=day_obj,
        )
        session.add(obj)

    obj.impression_to_order = conv.get("impression_to_order")
    obj.search_cat_to_cart  = conv.get("search_cat_to_cart")
    obj.search_cat_to_card  = conv.get("search_cat_to_card")
    obj.card_to_cart        = conv.get("card_to_cart")
    obj.cart_total          = conv.get("cart_total")
    obj.cart_to_order       = conv.get("cart_to_order")
    obj.order_to_purchase   = conv.get("order_to_purchase")

    obj.updated_at = datetime.utcnow()