# app/routers/rnp_import_reports_only.py
from __future__ import annotations

import asyncio
import re
import traceback
from uuid import uuid4
from io import BytesIO
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import pandas as pd
import openpyxl
from fastapi import APIRouter, Request, UploadFile, File
from fastapi.responses import JSONResponse
from sqlalchemy import select

from app.db import SessionLocal
from app.excel_repair import repair_xlsx_bytes
from app.ozon_performance_api import get_perf_client, OzonPerformanceApiError

from app.models import (
    Product,
    TrackedCampaign,
    AdsTopDaily,
    AdsStencilDaily,
    AdsTotalDaily,
    ConversionsDaily,
)

router = APIRouter()

# ---- in-memory jobs (dev/simple вариант)
BIDS_JOBS: dict[str, dict[str, Any]] = {}
BIDS_JOBS_LOCK = asyncio.Lock()

# ---- promo (аналитика продвижения/продаж) ожидаемые значения
INSTR_PPC = "Оплата за клик"
PLACE_TOP = "Поиск"
PLACE_TRAF = "Поиск и рекомендации"

# ---- Ozon placements для фильтра кампаний (для ставки конкурента)
PLACEMENT_TOP = "PLACEMENT_TOP_PROMOTION"
PLACEMENT_TRAF_SET = {
    "PLACEMENT_SEARCH_AND_CATEGORY",
    "PLACEMENT_SEARCH_PROMOTION",
    "PLACEMENT_STENCIL",
    "PLACEMENT_TRAFFIC",
}

BID_SCALE = Decimal("1000000")  # micro-rub -> rub


# =========================
# базовые парсеры/утилиты
# =========================
def _parse_ru_decimal(v: Any) -> Decimal:
    if v is None:
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))

    s = str(v).strip()
    if not s:
        return Decimal("0")
    s = s.replace("₽", "").replace("\u00a0", " ").replace("\u202f", " ").replace(" ", "")
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]+", "", s)
    if not s:
        return Decimal("0")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _parse_int(v: Any) -> int:
    if v is None:
        return 0
    try:
        s = str(v).strip().replace("\u00a0", " ").replace(" ", "")
        if not s:
            return 0
        if "," in s or "." in s:
            return int(_parse_ru_decimal(s))
        return int(s)
    except Exception:
        return 0


def _safe_div(a: Decimal, b: Decimal) -> Decimal | None:
    if b is None or b == 0:
        return None
    return a / b


def _bid_to_rub(v: Any) -> Decimal | None:
    if v is None:
        return None
    d = _parse_ru_decimal(v)
    return (d / BID_SCALE).quantize(Decimal("0.01"))


def _norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).replace("\u00a0", " ").strip() for c in df.columns]
    return df


def _extract_period_from_promo_bytes(promo_bytes: bytes) -> date | None:
    wb = openpyxl.load_workbook(BytesIO(promo_bytes), data_only=True)
    ws = wb["Statistics"] if "Statistics" in wb.sheetnames else wb.active
    v = ws["A1"].value or ""
    s = str(v)
    m = re.search(r"(\d{2}\.\d{2}\.\d{4})", s)
    if not m:
        return None
    return datetime.strptime(m.group(1), "%d.%m.%Y").date()


def _extract_period_from_sales_ws(ws) -> date | None:
    # у тебя: A2 "Период с dd.mm.yyyy по dd.mm.yyyy"
    v = ws.cell(2, 1).value
    if not v:
        return None
    m = re.search(r"(\d{2}\.\d{2}\.\d{4})", str(v))
    if not m:
        return None
    return datetime.strptime(m.group(1), "%d.%m.%Y").date()


def _norm_header(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def _find_header_row_by_title(ws, needle: str, max_scan_rows: int = 80, max_scan_cols: int = 80) -> int | None:
    n = _norm_header(needle)
    for r in range(1, max_scan_rows + 1):
        for c in range(1, max_scan_cols + 1):
            if _norm_header(ws.cell(r, c).value) == n:
                return r
    return None


def _build_sales_header_to_col(ws, header_row: int) -> dict[str, int]:
    header_to_col: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        top = ws.cell(header_row, c).value
        sub = ws.cell(header_row + 1, c).value
        name = sub if (sub is not None and str(sub).strip() != "") else top
        key = _norm_header(name)
        if key:
            header_to_col[key] = c
    return header_to_col


def _parse_excel_fraction_to_percent(v: Any) -> Decimal | None:
    if v is None:
        return None
    d = _parse_ru_decimal(v)
    if Decimal("0") <= d <= Decimal("1"):
        d = d * Decimal("100")
    return d


# =========================
# парсинг 2 отчётов
# =========================
def parse_top_traf_total_from_promo_report(promo_bytes: bytes, target_sku: str) -> dict[str, Any]:
    """
    Берём ТОП/ТРАФ/ИТОГО из promo-отчёта (Statistics):
    - фильтрация по SKU
    - ТОП: Инструмент=Оплата за клик AND Место размещения=Поиск
    - ТРАФ: ... = Поиск и рекомендации
    - ИТОГО: все строки по SKU, always_weighted=True
    """
    day_obj = _extract_period_from_promo_bytes(promo_bytes)

    df = pd.read_excel(BytesIO(promo_bytes), sheet_name="Statistics", engine="openpyxl", header=1)
    df = _norm_cols(df)

    if "SKU" not in df.columns:
        raise ValueError("В promo-отчёте нет колонки SKU")

    df["SKU"] = df["SKU"].astype(str).str.strip()
    for c in ("Инструмент", "Место размещения"):
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip()

    df_sku = df[df["SKU"] == str(target_sku)]
    if df_sku.empty:
        raise ValueError(f"В promo-отчёте нет строк по SKU={target_sku}")

    def agg(df_part: pd.DataFrame, *, always_weighted: bool) -> dict[str, Any]:
        df_part = _norm_cols(df_part)
        if df_part.empty:
            return {
                "spend_rub": Decimal("0"),
                "sales_rub": Decimal("0"),
                "orders_cnt": 0,
                "impressions": 0,
                "clicks": 0,
                "carts": 0,
                "drr": None,
                "ctr": None,
                "conv": None,
                "cpc": None,
                "cpo": None,
            }

        spend = sum((_parse_ru_decimal(x) for x in df_part["Расход, ₽"].tolist()), Decimal("0"))
        sales = sum((_parse_ru_decimal(x) for x in df_part["Продажи, ₽"].tolist()), Decimal("0"))
        orders = int(sum((_parse_ru_decimal(x) for x in df_part["Заказы, шт"].tolist()), Decimal("0")))
        impressions = int(sum((_parse_ru_decimal(x) for x in df_part["Показы"].tolist()), Decimal("0")))
        clicks = int(sum((_parse_ru_decimal(x) for x in df_part["Клики"].tolist()), Decimal("0")))
        carts = int(sum((_parse_ru_decimal(x) for x in df_part["В корзину"].tolist()), Decimal("0")))

        # если 1 строка и не always_weighted — берём готовые проценты (как в твоём main.py)
        if (len(df_part.index) == 1) and (not always_weighted):
            row = df_part.iloc[0]
            drr = _parse_ru_decimal(row.get("ДРР, %"))
            ctr = _parse_ru_decimal(row.get("CTR, %"))
            conv = _parse_ru_decimal(row.get("Конверсия в корзину, %"))
            cpo = _parse_ru_decimal(row.get("Затраты на заказ, ₽"))
            cpc = _parse_ru_decimal(row.get("Стоимость клика, ₽"))
        else:
            drr = (spend / sales * Decimal("100")) if sales > 0 else Decimal("0")
            ctr = (Decimal(clicks) / Decimal(impressions) * Decimal("100")) if impressions > 0 else Decimal("0")
            conv = (Decimal(carts) / Decimal(clicks) * Decimal("100")) if clicks > 0 else Decimal("0")
            cpc = (spend / Decimal(clicks)) if clicks > 0 else Decimal("0")
            cpo = (spend / Decimal(orders)) if orders > 0 else Decimal("0")

        return {
            "spend_rub": spend,
            "sales_rub": sales,
            "orders_cnt": orders,
            "impressions": impressions,
            "clicks": clicks,
            "carts": carts,
            "drr": drr,
            "ctr": ctr,
            "conv": conv,
            "cpc": cpc,  # ✅ “наша ставка” = средняя стоимость клика из отчёта
            "cpo": cpo,
        }

    df_top = df_sku[(df_sku.get("Инструмент") == INSTR_PPC) & (df_sku.get("Место размещения") == PLACE_TOP)]
    df_traf = df_sku[(df_sku.get("Инструмент") == INSTR_PPC) & (df_sku.get("Место размещения") == PLACE_TRAF)]

    top = agg(df_top, always_weighted=False)
    traf = agg(df_traf, always_weighted=False)
    total = agg(df_sku, always_weighted=True)

    return {
        "day_obj": day_obj,
        "top": top,
        "traf": traf,
        "total": total,
        "campaigns_in_file_for_sku": int(len(df_sku.index)),
    }


def parse_orders_and_conversions_from_sales_report(sales_fixed_bytes: bytes, target_sku: str) -> dict[str, Any]:
    wb = openpyxl.load_workbook(BytesIO(sales_fixed_bytes), data_only=True)
    ws = wb["По товарам"] if "По товарам" in wb.sheetnames else wb.active

    day_obj = _extract_period_from_sales_ws(ws)

    header_row = _find_header_row_by_title(ws, "SKU")
    if not header_row:
        raise ValueError("В sales-отчёте не нашёл заголовок 'SKU'")

    header_to_col = _build_sales_header_to_col(ws, header_row)

    need_orders = {
        "sku": "sku",
        "заказано на сумму": "all_orders_amount_rub",
        "заказано товаров": "all_orders_count",
    }
    for k in need_orders.keys():
        if k not in header_to_col:
            raise ValueError(f"В sales-отчёте нет колонки: {k}")

    need_conv = {
        "конверсия из показа в заказ": "impression_to_order",
        "конверсия из поиска и каталога в корзину": "search_cat_to_cart",
        "конверсия из поиска и каталога в карточку": "search_cat_to_card",
        "конверсия из карточки в корзину": "card_to_cart",
        "конверсия в корзину общая": "cart_total",
        "конверсия из корзины в заказ": "cart_to_order",
        "конверсия из заказа в выкуп": "order_to_purchase",
    }
    for k in need_conv.keys():
        if k not in header_to_col:
            raise ValueError(f"В sales-отчёте нет колонки конверсии: {k}")

    col_sku = header_to_col["sku"]

    found_row = None
    r = header_row + 1
    empty_streak = 0
    while r <= ws.max_row:
        sku_val = ws.cell(r, col_sku).value
        if sku_val is None or str(sku_val).strip() == "":
            empty_streak += 1
            if empty_streak >= 30:
                break
            r += 1
            continue
        empty_streak = 0

        try:
            sku_int = str(int(str(sku_val).strip()))
        except Exception:
            r += 1
            continue

        if sku_int == str(int(target_sku)):
            found_row = r
            break
        r += 1

    if not found_row:
        raise ValueError(f"В sales-отчёте не нашёл строку по SKU={target_sku}")

    all_orders_amount = _parse_ru_decimal(ws.cell(found_row, header_to_col["заказано на сумму"]).value)
    all_orders_count = _parse_int(ws.cell(found_row, header_to_col["заказано товаров"]).value)

    conv: dict[str, Decimal | None] = {}
    for title_norm, field in need_conv.items():
        v = ws.cell(found_row, header_to_col[title_norm]).value
        conv[field] = _parse_excel_fraction_to_percent(v)

    return {
        "day_obj": day_obj,
        "all_orders_amount_rub": all_orders_amount,
        "all_orders_count": all_orders_count,
        "conversions": conv,
    }


# =========================
# конкурентная ставка из API (только она)
# =========================
async def _get_active_product_and_sku(request: Request) -> tuple[int, int, str]:
    ozon_account_id = request.session.get("active_ozon_account_id")
    if not ozon_account_id:
        raise OzonPerformanceApiError("Не выбран Ozon-кабинет")

    product_id = request.session.get("active_product_id")
    if not product_id:
        raise OzonPerformanceApiError("Не выбран товар на РНП")

    async with SessionLocal() as session:
        product = await session.get(Product, int(product_id))

    if not product:
        raise OzonPerformanceApiError(f"Товар product_id={product_id} не найден")
    if int(product.ozon_account_id) != int(ozon_account_id):
        raise OzonPerformanceApiError("Товар не принадлежит выбранному кабинету")
    if product.sku is None:
        raise OzonPerformanceApiError("У товара не заполнен sku")

    return int(ozon_account_id), int(product.product_id), str(int(product.sku))


async def _get_tracked_campaign_ids(ozon_account_id: int) -> set[str]:
    async with SessionLocal() as session:
        rows = await session.execute(
            select(TrackedCampaign.campaign_id).where(TrackedCampaign.ozon_account_id == int(ozon_account_id))
        )
        return {str(x) for x in rows.scalars().all() if x}


def _has_any_placement(c: dict[str, Any], need: str | set[str]) -> bool:
    placements = c.get("placement") or []
    if isinstance(need, set):
        return any(p in need for p in placements)
    return need in placements


async def _fetch_competitor_bid_rub(api, *, campaign: dict[str, Any], target_sku: str) -> Decimal | None:
    """
    Берём конкурентную ставку по SKU в кампании (competitive endpoint).

    ВАЖНО: сознательно убрали предварительный запрос campaign_products_v2_all —
    это резко ускоряет, потому что иначе на каждую кампанию было 2 запроса.
    Если конкурентный endpoint вернёт пусто/ошибку — просто считаем, что ставки нет.
    """
    campaign_id = str(campaign.get("id"))

    competitive = await api.campaign_products_bids_competitive_all(
        campaign_id=campaign_id,
        skus=[target_sku],
        chunk_size=200,
    )
    bids = (competitive or {}).get("bids") or []
    if not bids:
        return None
    bid_micro = (bids[0] or {}).get("bid")
    return _bid_to_rub(bid_micro)


async def _gather_max_bid(api, campaigns: list[dict[str, Any]], target_sku: str, *, limit: int = 5) -> Decimal | None:
    sem = asyncio.Semaphore(limit)

    async def one(c):
        async with sem:
            return await _fetch_competitor_bid_rub(api, campaign=c, target_sku=target_sku)

    res = await asyncio.gather(*(one(c) for c in campaigns), return_exceptions=True)
    bids: list[Decimal] = []
    for x in res:
        if isinstance(x, Exception):
            continue
        if x is not None:
            bids.append(x)
    return max(bids) if bids else None


async def get_competitor_bids_only(request: Request, target_sku: str) -> dict[str, Decimal | None]:
    """
    Возвращаем max конкурентную ставку по tracked-кампаниям отдельно для ТОП и ТРАФ.
    """
    api = await get_perf_client(request)
    ozon_account_id = request.session.get("active_ozon_account_id")
    tracked_ids = await _get_tracked_campaign_ids(int(ozon_account_id))

    campaigns_data = await api.campaign_list()
    campaigns_all = (campaigns_data.get("list", []) or []) if isinstance(campaigns_data, dict) else []
    campaigns = [c for c in campaigns_all if str(c.get("id")) in tracked_ids]

    top_campaigns = [c for c in campaigns if _has_any_placement(c, PLACEMENT_TOP)]
    traf_campaigns = [c for c in campaigns if _has_any_placement(c, PLACEMENT_TRAF_SET)]

    top_max = await _gather_max_bid(api, top_campaigns, target_sku, limit=5)
    traf_max = await _gather_max_bid(api, traf_campaigns, target_sku, limit=5)

    return {"top": top_max, "traf": traf_max}


# =========================
# UPSERT (метрики из отчётов)
# =========================
async def _upsert_ads_top_daily_from_report(
    *,
    session,
    ozon_account_id: int,
    product_id: int,
    day_obj: date,
    m: dict[str, Any],
) -> None:
    q = await session.execute(
        select(AdsTopDaily).where(
            AdsTopDaily.ozon_account_id == ozon_account_id,
            AdsTopDaily.product_id == product_id,
            AdsTopDaily.day == day_obj,
        )
    )
    obj = q.scalar_one_or_none()
    if obj is None:
        obj = AdsTopDaily(ozon_account_id=ozon_account_id, product_id=product_id, day=day_obj)
        session.add(obj)

    obj.spend_rub = m["spend_rub"].quantize(Decimal("0.01"))
    obj.orders_amount_rub = m["sales_rub"].quantize(Decimal("0.01"))
    obj.orders_count = int(m["orders_cnt"])
    obj.impressions = int(m["impressions"])
    obj.clicks = int(m["clicks"])
    obj.carts = int(m["carts"])
    obj.ctr = (m["ctr"].quantize(Decimal("0.0001")) if m["ctr"] is not None else None)
    obj.drr = (m["drr"].quantize(Decimal("0.0001")) if m["drr"] is not None else None)

    # ✅ “наша ставка” = средняя стоимость клика из отчёта (CPC)
    obj.bid_ours = (m["cpc"].quantize(Decimal("0.01")) if m.get("cpc") is not None else None)

    # производные (как у тебя)
    obj.spend_per_order = (
        _safe_div(obj.spend_rub, Decimal(obj.orders_count))
        if obj.orders_count
        else Decimal("0")
    )
    obj.ad_conversion = (
        _safe_div(Decimal(obj.carts) * Decimal("100"), Decimal(obj.clicks)) if obj.clicks else Decimal("0")
    )

    obj.updated_at = datetime.utcnow()


async def _upsert_ads_stencil_daily_from_report(
    *,
    session,
    ozon_account_id: int,
    product_id: int,
    day_obj: date,
    m: dict[str, Any],
) -> None:
    q = await session.execute(
        select(AdsStencilDaily).where(
            AdsStencilDaily.ozon_account_id == ozon_account_id,
            AdsStencilDaily.product_id == product_id,
            AdsStencilDaily.day == day_obj,
        )
    )
    obj = q.scalar_one_or_none()
    if obj is None:
        obj = AdsStencilDaily(ozon_account_id=ozon_account_id, product_id=product_id, day=day_obj)
        session.add(obj)

    obj.spend_rub = m["spend_rub"].quantize(Decimal("0.01"))
    obj.orders_amount_rub = m["sales_rub"].quantize(Decimal("0.01"))
    obj.orders_count = int(m["orders_cnt"])
    obj.impressions = int(m["impressions"])
    obj.clicks = int(m["clicks"])
    obj.carts = int(m["carts"])
    obj.ctr = (m["ctr"].quantize(Decimal("0.0001")) if m["ctr"] is not None else None)
    obj.drr = (m["drr"].quantize(Decimal("0.0001")) if m["drr"] is not None else None)

    obj.bid_ours = (m["cpc"].quantize(Decimal("0.01")) if m.get("cpc") is not None else None)

    obj.spend_per_order = (
        _safe_div(obj.spend_rub, Decimal(obj.orders_count))
        if obj.orders_count
        else Decimal("0")
    )
    obj.ad_conversion = (
        _safe_div(Decimal(obj.carts) * Decimal("100"), Decimal(obj.clicks)) if obj.clicks else Decimal("0")
    )

    obj.updated_at = datetime.utcnow()


async def _upsert_ads_total_daily_from_report(
    *,
    session,
    ozon_account_id: int,
    product_id: int,
    day_obj: date,
    total: dict[str, Any],
    all_orders_amount_rub: Decimal,
    all_orders_count: int,
) -> None:
    q = await session.execute(
        select(AdsTotalDaily).where(
            AdsTotalDaily.ozon_account_id == ozon_account_id,
            AdsTotalDaily.product_id == product_id,
            AdsTotalDaily.day == day_obj,
        )
    )
    obj = q.scalar_one_or_none()
    if obj is None:
        obj = AdsTotalDaily(ozon_account_id=ozon_account_id, product_id=product_id, day=day_obj)
        session.add(obj)

    obj.spend_rub = total["spend_rub"].quantize(Decimal("0.01"))
    obj.ad_orders_amount_rub = total["sales_rub"].quantize(Decimal("0.01"))
    obj.ad_orders_count = int(total["orders_cnt"])
    obj.impressions = int(total["impressions"])
    obj.clicks = int(total["clicks"])
    obj.carts = int(total["carts"])
    obj.ctr = (total["ctr"].quantize(Decimal("0.0001")) if total["ctr"] is not None else None)
    obj.drr = (total["drr"].quantize(Decimal("0.0001")) if total["drr"] is not None else None)
    obj.ad_conversion = (total["conv"].quantize(Decimal("0.0001")) if total["conv"] is not None else None)

    obj.all_orders_amount_rub = all_orders_amount_rub.quantize(Decimal("0.01"))
    obj.all_orders_count = int(all_orders_count)

    # общий ДРР = расход / все заказы
    obj.drr_total = (obj.spend_rub / obj.all_orders_amount_rub * Decimal("100")) if obj.all_orders_amount_rub else None

    obj.updated_at = datetime.utcnow()


async def _upsert_conversions_daily_from_report(
    *,
    session,
    ozon_account_id: int,
    product_id: int,
    day_obj: date,
    conv: dict[str, Decimal | None],
) -> None:
    q = await session.execute(
        select(ConversionsDaily).where(
            ConversionsDaily.ozon_account_id == ozon_account_id,
            ConversionsDaily.product_id == product_id,
            ConversionsDaily.day == day_obj,
        )
    )
    obj = q.scalar_one_or_none()
    if obj is None:
        obj = ConversionsDaily(ozon_account_id=ozon_account_id, product_id=product_id, day=day_obj)
        session.add(obj)

    obj.impression_to_order = conv.get("impression_to_order")
    obj.search_cat_to_cart = conv.get("search_cat_to_cart")
    obj.search_cat_to_card = conv.get("search_cat_to_card")
    obj.card_to_cart = conv.get("card_to_cart")
    obj.cart_total = conv.get("cart_total")
    obj.cart_to_order = conv.get("cart_to_order")
    obj.order_to_purchase = conv.get("order_to_purchase")

    obj.updated_at = datetime.utcnow()


async def _upsert_bids_only(
    *,
    session,
    model,
    ozon_account_id: int,
    product_id: int,
    day_obj: date,
    bid_comp: Decimal | None,
) -> None:
    q = await session.execute(
        select(model).where(
            model.ozon_account_id == ozon_account_id,
            model.product_id == product_id,
            model.day == day_obj,
        )
    )
    obj = q.scalar_one_or_none()
    if obj is None:
        obj = model(ozon_account_id=ozon_account_id, product_id=product_id, day=day_obj)
        session.add(obj)

    if bid_comp is not None:
        obj.bid_competitor = bid_comp.quantize(Decimal("0.01"))

    obj.updated_at = datetime.utcnow()


# =========================
# background task: ставки
# =========================
async def _bg_fetch_bids_and_save(
    *,
    job_id: str,
    request: Request,
    ozon_account_id: int,
    product_id: int,
    target_sku: str,
    day_report: date,
    bid_day: date
) -> None:
    try:
        bids = await get_competitor_bids_only(request, target_sku=target_sku)

        async with SessionLocal() as session:
            # пишем competitor bids в ДЕНЬ ОТЧЁТА (day_obj)
            await _upsert_bids_only(
                session=session,
                model=AdsTopDaily,
                ozon_account_id=ozon_account_id,
                product_id=product_id,
                day_obj=bid_day,
                bid_comp=bids["top"],
            )
            await _upsert_bids_only(
                session=session,
                model=AdsStencilDaily,
                ozon_account_id=ozon_account_id,
                product_id=product_id,
                day_obj=bid_day,
                bid_comp=bids["traf"],
            )
            await session.commit()

        async with BIDS_JOBS_LOCK:
            job = BIDS_JOBS.get(job_id)
            if job is not None:
                job["status"] = "done"
                job["bids"] = {
                    "top": str(bids["top"]) if bids["top"] is not None else None,
                    "traf": str(bids["traf"]) if bids["traf"] is not None else None,
                }
                job["finished_at"] = datetime.utcnow().isoformat()

    except Exception as e:
        async with BIDS_JOBS_LOCK:
            job = BIDS_JOBS.get(job_id)
            if job is not None:
                job["status"] = "error"
                job["error"] = str(e)
                job["finished_at"] = datetime.utcnow().isoformat()


# =========================
# ENDPOINTS
# =========================
@router.get("/api/rnp/bids-job/{job_id}")
async def rnp_bids_job_status(job_id: str):
    async with BIDS_JOBS_LOCK:
        job = BIDS_JOBS.get(job_id)
    if not job:
        return JSONResponse(status_code=404, content={"ok": False, "error": "job not found"})
    return JSONResponse(content={"ok": True, "job": job})


@router.post("/api/rnp/import-reports")
async def rnp_import_reports(
    request: Request,
    promo_report: UploadFile = File(...),  # отчёт аналитики продаж/продвижения (Statistics)
    sales_report: UploadFile = File(...),  # отчёт продаж товара (битый)
):
    try:
        ozon_account_id, product_id, target_sku = await _get_active_product_and_sku(request)

        promo_bytes = await promo_report.read()
        sales_bytes = await sales_report.read()
        sales_fixed = repair_xlsx_bytes(sales_bytes)

        promo_parsed = parse_top_traf_total_from_promo_report(promo_bytes, target_sku)
        sales_parsed = parse_orders_and_conversions_from_sales_report(sales_fixed, target_sku)

        day_promo = promo_parsed["day_obj"]
        day_sales = sales_parsed["day_obj"]

        # день отчёта — берём тот, который нашёлся; если оба есть и разные — лучше падать, чтобы не мешать данные
        day_obj = day_promo or day_sales
        if not day_obj:
            raise ValueError("Не смог определить дату отчёта (ни promo A1, ни sales A2)")
        if day_promo and day_sales and (day_promo != day_sales):
            raise ValueError(f"Даты в отчётах не совпадают: promo={day_promo} sales={day_sales}")

        async with SessionLocal() as session:
            await _upsert_ads_top_daily_from_report(
                session=session,
                ozon_account_id=ozon_account_id,
                product_id=product_id,
                day_obj=day_obj,
                m=promo_parsed["top"],
            )
            await _upsert_ads_stencil_daily_from_report(
                session=session,
                ozon_account_id=ozon_account_id,
                product_id=product_id,
                day_obj=day_obj,
                m=promo_parsed["traf"],
            )

            await _upsert_ads_total_daily_from_report(
                session=session,
                ozon_account_id=ozon_account_id,
                product_id=product_id,
                day_obj=day_obj,
                total=promo_parsed["total"],
                all_orders_amount_rub=sales_parsed["all_orders_amount_rub"],
                all_orders_count=sales_parsed["all_orders_count"],
            )

            await _upsert_conversions_daily_from_report(
                session=session,
                ozon_account_id=ozon_account_id,
                product_id=product_id,
                day_obj=day_obj,
                conv=sales_parsed["conversions"],
            )

            await session.commit()

        bid_day_obj = date.today()
        # ---- запускаем ставки в фоне (НЕ ждём)
        job_id = str(uuid4())
        async with BIDS_JOBS_LOCK:
            BIDS_JOBS[job_id] = {
                "status": "pending",
                "created_at": datetime.utcnow().isoformat(),
                "ozon_account_id": ozon_account_id,
                "product_id": product_id,
                "sku": target_sku,
                "day_report": day_obj.isoformat(),
                "day_bids": bid_day_obj.isoformat(),
            }

        asyncio.create_task(
            _bg_fetch_bids_and_save(
                job_id=job_id,
                request=request,
                ozon_account_id=ozon_account_id,
                product_id=product_id,
                target_sku=target_sku,
                day_report=day_obj,
                bid_day=bid_day_obj,
            )
        )

        return JSONResponse(
            content={
                "ok": True,
                "message": "Импорт выполнен. Ставки конкурентов загружаются в фоне — можно закрыть окно.",
                "day_report": day_obj.isoformat(),
                "campaigns_in_file_for_sku": promo_parsed["campaigns_in_file_for_sku"],
                "bids_job_id": job_id,
            }
        )

    except OzonPerformanceApiError as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})
    except Exception:
        tb = traceback.format_exc()
        print(tb)  # ✅ будет в консоли uvicorn
        return JSONResponse(
            status_code=500,
            content={"ok": False, "error": "Unexpected error:\n" + tb},
        )
